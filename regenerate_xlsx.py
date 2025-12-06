#!/usr/bin/env python3
"""
Regenerate .xlsx stats files from gameshistory.json data.
"""

import json
import os
from openpyxl import Workbook

def regenerate_xlsx(game_data, output_path):
    """Regenerate an xlsx file from parsed game data."""
    wb = Workbook()

    # Sheet 1: Game Details
    ws_details = wb.active
    ws_details.title = 'Game Details'
    details = game_data.get('details', {})
    headers = ['Game Type', 'Variant Name', 'Map Name', 'Start Time', 'End Time', 'Duration']
    ws_details.append(headers)
    ws_details.append([
        details.get('Game Type', ''),
        details.get('Variant Name', ''),
        details.get('Map Name', ''),
        details.get('Start Time', ''),
        details.get('End Time', ''),
        details.get('Duration', '')
    ])

    # Sheet 2: Post Game Report
    ws_post = wb.create_sheet('Post Game Report')
    players = game_data.get('players', [])
    if players:
        post_headers = ['name', 'place', 'score', 'kills', 'deaths', 'assists', 'kda', 'suicides', 'team', 'shots_fired', 'shots_hit', 'accuracy', 'head_shots']
        ws_post.append(post_headers)
        for p in players:
            ws_post.append([
                p.get('name', ''),
                p.get('place', ''),
                p.get('score', 0),
                p.get('kills', 0),
                p.get('deaths', 0),
                p.get('assists', 0),
                p.get('kda', 0),
                p.get('suicides', 0),
                p.get('team', ''),
                p.get('shots_fired', 0),
                p.get('shots_hit', 0),
                p.get('accuracy', 0),
                p.get('head_shots', 0)
            ])

    # Sheet 3: Versus (kill matrix)
    ws_versus = wb.create_sheet('Versus')
    versus = game_data.get('versus', {})
    if versus:
        player_names = list(versus.keys())
        # Header row: empty + all player names
        ws_versus.append([''] + player_names)
        for player in player_names:
            row = [player]
            for opponent in player_names:
                row.append(versus.get(player, {}).get(opponent, 0))
            ws_versus.append(row)

    # Sheet 4: Game Statistics
    ws_stats = wb.create_sheet('Game Statistics')
    detailed = game_data.get('detailed_stats', [])
    if detailed:
        stat_headers = ['Player', 'Emblem URL', 'kills', 'assists', 'deaths', 'headshots', 'betrayals', 'suicides', 'best_spree', 'total_time_alive', 'ctf_scores', 'ctf_flag_steals', 'ctf_flag_saves']
        ws_stats.append(stat_headers)
        for s in detailed:
            ws_stats.append([
                s.get('player', ''),
                s.get('emblem_url', ''),
                s.get('kills', 0),
                s.get('assists', 0),
                s.get('deaths', 0),
                s.get('headshots', 0),
                s.get('betrayals', 0),
                s.get('suicides', 0),
                s.get('best_spree', 0),
                s.get('total_time_alive', 0),
                s.get('ctf_scores', 0),
                s.get('ctf_flag_steals', 0),
                s.get('ctf_flag_saves', 0)
            ])

    # Sheet 5: Medal Stats
    ws_medals = wb.create_sheet('Medal Stats')
    medals = game_data.get('medals', [])
    if medals:
        medal_cols = ['player', 'double_kill', 'triple_kill', 'killtacular', 'kill_frenzy', 'killtrocity',
                     'killamanjaro', 'sniper_kill', 'road_kill', 'bone_cracker', 'assassin',
                     'vehicle_destroyed', 'car_jacking', 'stick_it', 'killing_spree',
                     'running_riot', 'rampage', 'beserker', 'over_kill', 'flag_taken',
                     'flag_carrier_kill', 'flag_returned', 'bomb_planted', 'bomb_carrier_kill', 'bomb_returned']
        ws_medals.append(medal_cols)
        for m in medals:
            row = [m.get('player', '')]
            for col in medal_cols[1:]:
                row.append(m.get(col, 0))
            ws_medals.append(row)

    # Sheet 6: Weapon Statistics
    ws_weapons = wb.create_sheet('Weapon Statistics')
    weapons = game_data.get('weapons', [])
    if weapons:
        # Get all weapon columns from first player
        weapon_cols = list(weapons[0].keys()) if weapons else ['Player']
        ws_weapons.append(weapon_cols)
        for w in weapons:
            row = [w.get(col, 0) for col in weapon_cols]
            ws_weapons.append(row)

    wb.save(output_path)
    print(f"  Created: {output_path}")

def main():
    # Load gameshistory.json
    with open('/home/user/CarnageReport.com/gameshistory.json') as f:
        games = json.load(f)

    # Find 12/2 games and regenerate them
    output_dir = '/home/user/CarnageReport.com/stats'
    os.makedirs(output_dir, exist_ok=True)

    count = 0
    for game in games:
        source_file = game.get('source_file', '')
        if '20251202' in source_file:
            output_path = os.path.join(output_dir, source_file)
            regenerate_xlsx(game, output_path)
            count += 1

    print(f"\nRegenerated {count} files")

if __name__ == '__main__':
    main()
